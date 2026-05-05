"""
RMSD to Prism
==============
Computes Cα RMSD over time for each simulation and outputs
a single wide-format Excel table ready for GraphPad Prism:

    Time (ns) | sim_0 | sim_1 | sim_2 | ...

Usage:
    python rmsd_to_prism.py --input /path/to/folder/

    # With selection:
    python rmsd_to_prism.py --input /path/to/folder/ --chain PROA --resid 50:200

    # All options:
    python rmsd_to_prism.py --input /path/to/folder/ \\
        --psf now_0000.psf \\
        --dcd-pattern "trajectory_now_{:d}.dcd" \\
        --n-sims 72 \\
        --chain PROA \\
        --resid 50:200 \\
        --output rmsd_prism.xlsx
"""

import os
import sys
import argparse
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import MDAnalysis as mda
from MDAnalysis.analysis import rms
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="RMSD → Prism: compute Cα RMSD for all sims, output wide table",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python rmsd_to_prism.py --input /data/sims/
  python rmsd_to_prism.py --input /data/sims/ --chains PROA --resid 50:200
  python rmsd_to_prism.py --input /data/sims/ --chains PROA PROD --resid 50:200 --output rmsd_dimer.xlsx
  python rmsd_to_prism.py --input /data/sims/ --chains PROA PROD --resid 50:100,150:300 --output rmsd_TM.xlsx
        """
    )
    parser.add_argument("--input", "-i", metavar="FOLDER", required=True,
        help="Folder containing PSF + DCD files.")
    parser.add_argument("--psf", metavar="FILENAME", default="now_0000.psf",
        help="PSF filename (default: now_0000.psf)")
    parser.add_argument("--dcd-pattern", metavar="PATTERN",
        default="trajectory_now_{:d}.dcd",
        help='DCD filename pattern (default: "trajectory_now_{:d}.dcd")')
    parser.add_argument("--n-sims", type=int, default=72, metavar="N",
        help="Number of simulations (default: 72)")
    parser.add_argument("--chains", nargs="+", metavar="SEGID", default=None,
        help="One or more chain/segid(s) for RMSD (e.g. PROA or PROA PROD). Omit for all protein.")
    parser.add_argument("--resid", metavar="RANGE", default=None,
        help="Residue range(s), e.g. '50:200' or '50:100,150:200'.")
    parser.add_argument("--output", "-o", metavar="FILE", default="rmsd_prism.xlsx",
        help="Output Excel filename (default: rmsd_prism.xlsx)")
    return parser.parse_args()

args = parse_args()

INPUT_FOLDER = args.input
PSF_FILENAME = args.psf
DCD_PATTERN  = args.dcd_pattern
N_SIMS       = args.n_sims
CHAINS       = args.chains
RESID_RANGE  = args.resid
OUTPUT_XLSX  = args.output

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def build_ca_selection():
    # Chain base
    if not CHAINS:
        base = "protein"
    elif len(CHAINS) == 1:
        base = f"segid {CHAINS[0]}"
    else:
        base = f"(segid {' or segid '.join(CHAINS)})"
    base += " and name CA"

    if not RESID_RANGE:
        return base

    # Parse residue ranges
    parts = []
    for seg in RESID_RANGE.split(","):
        seg = seg.strip()
        if ":" in seg:
            s, e = seg.split(":", 1)
            parts.append(f"resid {s.strip()}:{e.strip()}")
        else:
            parts.append(f"resid {seg}")
    joined = parts[0] if len(parts) == 1 else f"({' or '.join(parts)})"
    return f"{base} and {joined}"


def discover_simulations():
    if not os.path.isdir(INPUT_FOLDER):
        print(f"[ERROR] Folder not found: {INPUT_FOLDER}"); sys.exit(1)
    psf = os.path.join(INPUT_FOLDER, PSF_FILENAME)
    if not os.path.exists(psf):
        print(f"[ERROR] PSF not found: {psf}"); sys.exit(1)

    sims, missing = [], []
    for i in range(N_SIMS):
        dcd = os.path.join(INPUT_FOLDER, DCD_PATTERN.format(i))
        if not os.path.exists(dcd):
            missing.append(i); continue
        sims.append((f"sim_{i}", psf, dcd))

    if missing:
        print(f"  [WARN] Missing DCDs for sims: {missing}")
    print(f"  PSF : {PSF_FILENAME}")
    print(f"  DCDs: {len(sims)}/{N_SIMS} found")
    return sims


def compute_rmsd(psf, dcd, sel_str):
    u   = mda.Universe(psf, dcd)
    ca  = u.select_atoms(sel_str)
    if len(ca) == 0:
        raise ValueError(f"No atoms found with selection: '{sel_str}'")
    ref = u.copy(); ref.trajectory[0]
    R   = rms.RMSD(ca, ref.select_atoms(sel_str), select=sel_str)
    R.run()
    times_ns = R.results.rmsd[:, 1] / 1000.0
    rmsd     = R.results.rmsd[:, 2]
    return times_ns, rmsd

# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

def write_prism_excel(df, sel_str, path):
    CLR_HEADER = "1F4E79"
    CLR_ALT    = "D6E4F0"
    thin       = Side(style="thin", color="BFBFBF")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def s(cell, bold=False, fill=None, fg="000000", fmt=None):
        cell.font      = Font(name="Arial", size=10, bold=bold, color=fg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        if fill: cell.fill = PatternFill("solid", start_color=fill)
        if fmt:  cell.number_format = fmt

    wb = Workbook()
    ws = wb.active
    ws.title = "RMSD"
    ws.sheet_view.showGridLines = False

    n_cols = len(df.columns)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(n_cols, 30))
    c = ws["A1"]
    c.value = "Cα RMSD vs Time — Prism Export"
    c.font  = Font(name="Arial", size=13, bold=True, color=CLR_HEADER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(n_cols, 30))
    c = ws["A2"]
    c.value = f"Selection: {sel_str}   |   {n_cols - 1} simulations   |   RMSD in Å"
    c.font  = Font(name="Arial", size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    # Column headers
    for col, header in enumerate(df.columns, start=1):
        c = ws.cell(row=3, column=col, value=header)
        s(c, bold=True, fill=CLR_HEADER, fg="FFFFFF")
        ws.column_dimensions[get_column_letter(col)].width = 14 if col == 1 else 10
    ws.row_dimensions[3].height = 18

    # Data rows
    for i, row in enumerate(df.itertuples(index=False)):
        excel_row = 4 + i
        alt = CLR_ALT if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row, start=1):
            v = None if (isinstance(val, float) and np.isnan(val)) else val
            c = ws.cell(row=excel_row, column=col, value=v)
            s(c, fill=alt, fmt="0.0000" if col == 1 else "0.000")
        ws.row_dimensions[excel_row].height = 13

    wb.save(path)

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  RMSD → Prism")
    print("=" * 55)

    sims    = discover_simulations()
    sel_str = build_ca_selection()
    print(f"  Selection: {sel_str}\n")

    results, failed = {}, []
    ref_times = None

    for idx, (name, psf, dcd) in enumerate(sims):
        print(f"  [{idx + 1}/{len(sims)}] {name}", end=" ... ", flush=True)
        try:
            times, rmsd = compute_rmsd(psf, dcd, sel_str)
            results[name] = (times, rmsd)
            if ref_times is None or len(times) > len(ref_times):
                ref_times = times
            print(f"done ({len(times)} frames)")
        except Exception as e:
            print(f"ERROR: {e}")
            failed.append((name, str(e)))

    if not results:
        print("\n[ERROR] All simulations failed."); return

    # Build wide DataFrame
    df = pd.DataFrame({"Time (ns)": ref_times})
    for name, (times, rmsd) in results.items():
        if len(rmsd) < len(ref_times):
            rmsd = np.concatenate([rmsd, np.full(len(ref_times) - len(rmsd), np.nan)])
        df[name] = rmsd

    print(f"\nWriting: {OUTPUT_XLSX}")
    write_prism_excel(df, sel_str, OUTPUT_XLSX)

    if failed:
        print(f"\n[WARN] {len(failed)} simulation(s) failed:")
        for name, err in failed:
            print(f"  {name}: {err}")

    print(f"\n✓ {OUTPUT_XLSX}")
    print(f"  {len(results)} simulations  |  {len(ref_times)} frames  |  {len(df.columns)} columns")
    print("=" * 55)

if __name__ == "__main__":
    main()
