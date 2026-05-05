"""
Membrane Area Convergence Analysis Script
==========================================
Tracks total membrane area (X * Y box dimensions) over time
for multiple NAMD/CHARMM simulations using a single shared PSF
and one DCD per simulation.

Outputs:
  - Excel file: summary sheet (convergence matrix) + one sheet per simulation
  - PNG: membrane area time series overlay plot for all simulations

Usage:
    python membrane_area_convergence.py --input /path/to/folder/

    # All options:
    python membrane_area_convergence.py --input /path/to/folder/ \\
        --psf now_0000.psf \\
        --dcd-pattern "trajectory_now_{:d}.dcd" \\
        --n-sims 72 \\
        --area-threshold 5.0 \\
        --eq-fraction 0.5 \\
        --output membrane_area.xlsx
"""

import os
import sys
import argparse
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import MDAnalysis as mda
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Membrane Area Convergence Analysis — XY box area over time",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python membrane_area_convergence.py --input /data/simulations/
  python membrane_area_convergence.py --input /data/sims/ --area-threshold 3.0
  python membrane_area_convergence.py --input /data/sims/ --eq-fraction 0.33 --output membrane.xlsx
        """
    )
    parser.add_argument("--input", "-i", metavar="FOLDER", required=True,
        help="Folder containing PSF + DCD files.")
    parser.add_argument("--psf", metavar="FILENAME", default="now_0000.psf",
        help="PSF filename (default: now_0000.psf)")
    parser.add_argument("--dcd-pattern", metavar="PATTERN",
        default="trajectory_now_{:d}.dcd",
        help='DCD filename pattern (default: "trajectory_now_{:d}.dcd")')
    parser.add_argument("--n-sims", type=int, default=72, metavar="N",
        help="Number of DCD files / simulations (default: 72)")
    parser.add_argument("--area-threshold", type=float, default=5.0, metavar="PERCENT",
        help="CV (%%) of area in production window above this = FAIL (default: 5.0)")
    parser.add_argument("--eq-fraction", type=float, default=0.5, metavar="FRACTION",
        help="Fraction of trajectory used as production window (default: 0.5)")
    parser.add_argument("--output", "-o", metavar="FILE",
        default="membrane_area_convergence.xlsx",
        help="Output Excel filename (default: membrane_area_convergence.xlsx)")
    return parser.parse_args()

args = parse_args()

INPUT_FOLDER           = args.input
PSF_FILENAME           = args.psf
DCD_PATTERN            = args.dcd_pattern
N_SIMS                 = args.n_sims
AREA_CV_THRESHOLD      = args.area_threshold
EQUILIBRATION_FRACTION = args.eq_fraction
OUTPUT_XLSX            = args.output
OVERLAY_PNG            = os.path.splitext(OUTPUT_XLSX)[0] + "_overlay.png"

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────

CLR_HEADER_BG = "1F4E79"
CLR_SUBHDR_BG = "2E75B6"
CLR_ALT_ROW   = "D6E4F0"
CLR_PASS      = "C6EFCE"; CLR_PASS_FG = "276221"
CLR_WARN      = "FFEB9C"; CLR_WARN_FG = "7D5A00"
CLR_FAIL      = "FFC7CE"; CLR_FAIL_FG = "9C0006"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT_BODY   = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITLE  = Font(name="Arial", size=13, bold=True, color=CLR_HEADER_BG)

def style(cell, bold=False, fill=None, fg="000000", align="center", fmt=None):
    cell.font      = Font(name="Arial", size=10, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = BORDER
    if fill: cell.fill = PatternFill("solid", start_color=fill)
    if fmt:  cell.number_format = fmt

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def convergence_status(cv_pct):
    if cv_pct > AREA_CV_THRESHOLD:
        return CLR_FAIL, CLR_FAIL_FG, "FAIL"
    elif cv_pct > AREA_CV_THRESHOLD * 0.75:
        return CLR_WARN, CLR_WARN_FG, "WARN"
    else:
        return CLR_PASS, CLR_PASS_FG, "PASS"


def production_stats(values):
    """Mean, std, CV%, and drift of the production window (last eq_fraction)."""
    n     = len(values)
    start = int(n * (1 - EQUILIBRATION_FRACTION))
    prod  = values[start:]
    mean  = prod.mean()
    std   = prod.std()
    cv    = (std / mean * 100) if mean != 0 else 0.0
    mid   = len(prod) // 2
    drift = abs(prod[:mid].mean() - prod[mid:].mean()) if mid > 0 else 0.0
    return mean, std, cv, drift

# ─────────────────────────────────────────────
# DISCOVERY
# ─────────────────────────────────────────────

def discover_simulations():
    if not os.path.isdir(INPUT_FOLDER):
        print(f"[ERROR] Folder not found: {INPUT_FOLDER}"); sys.exit(1)

    psf = os.path.join(INPUT_FOLDER, PSF_FILENAME)
    if not os.path.exists(psf):
        print(f"[ERROR] PSF not found: {psf}")
        print(f"        Use --psf to specify a different filename."); sys.exit(1)

    sims, missing = [], []
    for i in range(N_SIMS):
        dcd = os.path.join(INPUT_FOLDER, DCD_PATTERN.format(i))
        if not os.path.exists(dcd):
            missing.append(DCD_PATTERN.format(i)); continue
        sims.append((f"sim_{i}", psf, dcd))

    if missing:
        print(f"  [WARN] {len(missing)} DCD(s) not found: {', '.join(missing[:5])}"
              + (" ..." if len(missing) > 5 else ""))
    print(f"  PSF : {PSF_FILENAME}")
    print(f"  DCDs: {len(sims)}/{N_SIMS} found")
    return sims

# ─────────────────────────────────────────────
# ANALYSIS
# ─────────────────────────────────────────────

def analyze_simulation(name, psf, dcd):
    u        = mda.Universe(psf, dcd)
    n_frames = len(u.trajectory)
    total_ns = n_frames * u.trajectory.dt / 1000.0

    times_ns = np.zeros(n_frames)
    area     = np.zeros(n_frames)

    for i, ts in enumerate(u.trajectory):
        times_ns[i] = ts.time / 1000.0
        area[i]     = ts.dimensions[0] * ts.dimensions[1]  # X * Y in Å²

    mean, std, cv, drift = production_stats(area)
    fill, fg, status     = convergence_status(cv)

    return {
        "name":      name,
        "n_frames":  n_frames,
        "total_ns":  total_ns,
        "times_ns":  times_ns,
        "area":      area,
        "mean":      mean,
        "std":       std,
        "cv":        cv,
        "drift":     drift,
        "fill":      fill,
        "fg":        fg,
        "status":    status,
    }

# ─────────────────────────────────────────────
# PLOTS
# ─────────────────────────────────────────────

def build_overlay_plot(results):
    """All simulation membrane area traces on one plot, colored by status."""
    fig, ax = plt.subplots(figsize=(12, 6))

    color_map = {"PASS": "#27AE60", "WARN": "#F39C12", "FAIL": "#E74C3C"}
    alpha_map = {"PASS": 0.4,       "WARN": 0.5,       "FAIL": 0.5}

    for status in ["PASS", "WARN", "FAIL"]:
        for r in results:
            if r["status"] != status: continue
            ax.plot(r["times_ns"], r["area"],
                    color=color_map[status], alpha=alpha_map[status], lw=0.8)

    # Legend proxies
    for status, label in [("PASS", "Converged"), ("WARN", "Warning"), ("FAIL", "Not Converged")]:
        n = sum(1 for r in results if r["status"] == status)
        if n:
            ax.plot([], [], color=color_map[status], lw=2, label=f"{label} (n={n})")

    # Production window shading
    if results:
        max_t = max(r["times_ns"][-1] for r in results)
        eq_t  = max_t * (1 - EQUILIBRATION_FRACTION)
        ax.axvspan(eq_t, max_t, alpha=0.06, color="#1F4E79", label="Production window")

    ax.set_xlabel("Time (ns)", fontsize=12)
    ax.set_ylabel("Membrane Area (Å²)", fontsize=12)
    ax.set_title(f"Membrane Area (XY) Time Series — {len(results)} Simulations",
                 fontsize=12, fontweight="bold", color="#1F4E79")
    ax.legend(framealpha=0.9, fontsize=9, loc="upper left")
    ax.grid(True, alpha=0.25, linestyle="--")
    ax.spines[["top", "right"]].set_visible(False)

    plt.tight_layout()
    plt.savefig(OVERLAY_PNG, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Overlay plot saved: {OVERLAY_PNG}")

# ─────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────

SUMMARY_COLS = [
    ("Simulation",         18),
    ("Length (ns)",        13),
    ("Frames",             10),
    ("Mean Area (Å²)",     15),
    ("Std Area (Å²)",      14),
    ("Area CV (%)",        12),
    ("Drift (Å²)",         12),
    ("Status",             10),
]

def build_summary_sheet(ws, results):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]; c.value = "Membrane Area Convergence Matrix"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (f"Metric: XY box area (Å²)   |   "
               f"Production window: last {int(EQUILIBRATION_FRACTION * 100)}%   |   "
               f"CV threshold: {AREA_CV_THRESHOLD}%")
    c.font = Font(name="Arial", size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 15

    # Column headers
    for col, (label, width) in enumerate(SUMMARY_COLS, start=1):
        c = ws.cell(row=3, column=col, value=label)
        style(c, bold=True, fill=CLR_HEADER_BG, fg="FFFFFF")
        col_width(ws, col, width)
    ws.row_dimensions[3].height = 18

    # Data rows
    for i, r in enumerate(results):
        row = 4 + i
        alt = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        vals = [r["name"], round(r["total_ns"], 2), r["n_frames"],
                round(r["mean"], 2), round(r["std"], 2),
                round(r["cv"], 3), round(r["drift"], 2),
                r["status"]]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 8:
                style(c, bold=True, fill=r["fill"], fg=r["fg"])
            else:
                style(c, fill=alt, align="left" if col == 1 else "center",
                      fmt=("0.00" if col in (4, 5, 7) else
                           "0.000" if col == 6 else None))
        ws.row_dimensions[row].height = 15

    # Average row
    n   = len(results)
    row = 4 + n
    ws.cell(row=row, column=1, value="AVERAGE")
    for col in range(1, 9):
        style(ws.cell(row=row, column=col), bold=True, fill="D9D9D9")
    for col, key, fmt in [(4, "mean", "0.00"), (5, "std", "0.00"),
                          (6, "cv",   "0.000"), (7, "drift", "0.00")]:
        c = ws.cell(row=row, column=col,
                    value=round(np.mean([r[key] for r in results]), 3))
        style(c, bold=True, fill="D9D9D9", fmt=fmt)
    ws.row_dimensions[row].height = 15


def build_overlay_sheet(ws):
    ws.title = "Overlay Plot"
    ws.sheet_view.showGridLines = False
    ws["A1"].value = "Membrane Area (XY) — All Simulations"
    ws["A1"].font  = FONT_TITLE
    if os.path.exists(OVERLAY_PNG):
        img = XLImage(OVERLAY_PNG)
        img.anchor = "A3"
        ws.add_image(img)


def build_timeseries_sheet(ws, r):
    ws.sheet_view.showGridLines = False

    ws["A1"].value = r["name"]
    ws["A1"].font  = Font(name="Arial", size=12, bold=True, color=CLR_HEADER_BG)

    # Stats box
    stats = [
        ("Metric",          "Value"),
        ("Length (ns)",     round(r["total_ns"], 2)),
        ("Frames",          r["n_frames"]),
        ("Mean Area (Å²)",  round(r["mean"],  2)),
        ("Std Area (Å²)",   round(r["std"],   2)),
        ("Area CV (%)",     round(r["cv"],    3)),
        ("Drift (Å²)",      round(r["drift"], 2)),
        ("Status",          r["status"]),
    ]
    for si, (label, val) in enumerate(stats):
        lc = ws.cell(row=2 + si, column=1, value=label)
        vc = ws.cell(row=2 + si, column=2, value=val)
        is_hdr = si == 0
        style(lc, bold=True,
              fill=CLR_HEADER_BG if is_hdr else "F2F2F2",
              fg="FFFFFF" if is_hdr else "000000")
        style(vc, bold=is_hdr or si == 7,
              fill=CLR_HEADER_BG if is_hdr else (r["fill"] if si == 7 else "F2F2F2"),
              fg="FFFFFF" if is_hdr else (r["fg"] if si == 7 else "000000"))

    # Time series table
    for col, label in [(4, "Time (ns)"), (5, "Area (Å²)")]:
        c = ws.cell(row=2, column=col, value=label)
        style(c, bold=True, fill=CLR_SUBHDR_BG, fg="FFFFFF")
        col_width(ws, col, 14)

    for i, (t, v) in enumerate(zip(r["times_ns"], r["area"])):
        alt = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        style(ws.cell(row=3 + i, column=4, value=round(float(t), 4)), fill=alt, fmt="0.0000")
        style(ws.cell(row=3 + i, column=5, value=round(float(v), 2)), fill=alt, fmt="0.00")

    col_width(ws, 1, 18); col_width(ws, 2, 14)

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  Membrane Area Convergence Analysis")
    print("=" * 55)

    sims = discover_simulations()
    print()

    results, failed = [], []
    for idx, (name, psf, dcd) in enumerate(sims):
        print(f"  [{idx + 1}/{len(sims)}] {name}", end=" ... ", flush=True)
        try:
            r = analyze_simulation(name, psf, dcd)
            results.append(r)
            print(f"mean={r['mean']:.1f} Å²  CV={r['cv']:.2f}%  [{r['status']}]")
        except Exception as e:
            print(f"ERROR: {e}")
            failed.append((name, str(e)))

    if not results:
        print("\n[ERROR] All simulations failed."); return

    print(f"\nCompleted: {len(results)}/{len(sims)} simulations")
    statuses = [r["status"] for r in results]
    print(f"  PASS: {statuses.count('PASS')}  WARN: {statuses.count('WARN')}  FAIL: {statuses.count('FAIL')}")

    print("\nGenerating overlay plot...")
    build_overlay_plot(results)

    print("Building Excel workbook...")
    wb = Workbook()

    build_summary_sheet(wb.active, results)

    ws_ov = wb.create_sheet("Overlay Plot")
    build_overlay_sheet(ws_ov)

    for r in results:
        ws = wb.create_sheet(r["name"][:31])
        build_timeseries_sheet(ws, r)

    if failed:
        ws_err = wb.create_sheet("Errors")
        ws_err["A1"].value = "Failed Simulations"
        ws_err["A1"].font  = Font(bold=True, color="FF0000")
        for i, (name, err) in enumerate(failed):
            ws_err.cell(row=2 + i, column=1, value=name)
            ws_err.cell(row=2 + i, column=2, value=err)

    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Excel saved : {OUTPUT_XLSX}")
    print(f"✓ Overlay plot: {OVERLAY_PNG}")
    print("=" * 55)

if __name__ == "__main__":
    main()
